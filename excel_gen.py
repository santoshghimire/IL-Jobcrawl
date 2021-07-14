import datetime
import pymysql
# import sys
# import locale
# import codecs
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import os

from jobcrawl.mailer import send_email

settings = {
    'MYSQL_HOST': 'localhost',
    'MYSQL_DBNAME': 'il_sites_datas',
    'MYSQL_USER': 'root',
    'MYSQL_PASSWORD': 'root'
}


def generate_excel(site_name):
    """ Read sql query (database table)  and return pandas dataframe"""
    # sys.stdout = codecs.getwriter(
    #     locale.getpreferredencoding())(sys.stdout)
    # reload(sys)
    # sys.setdefaultencoding('utf-8')

    today = datetime.date.today()
    today_str = today.strftime("%d/%m/%Y")
    site_name = site_name.lower().title()
    directory = 'IL-jobcrawl-data'
    file_name = '{}/{}_{}.xlsx'.format(
        directory, today.strftime("%Y_%m_%d"), site_name)
    main_writer = pd.ExcelWriter(
        file_name, engine='openpyxl')

    # today_str = '24/09/2016'
    names = {
        'Drushim': 'Drushim', 'Alljobs': 'AllJobs',
        'Jobmaster': 'JobMaster'}
    site = names.get(site_name)
    conn = pymysql.connect(
        host=settings.get('MYSQL_HOST'), port=3306,
        user=settings.get('MYSQL_USER'),
        passwd=settings.get('MYSQL_PASSWORD'),
        db=settings.get('MYSQL_DBNAME'),
        charset='utf8'
    )

    sql = """SELECT Site, Company, Company_jobs, Job_id, Job_title,
        Job_Description, Job_Post_Date, Job_URL, Country_Areas,
        Job_categories, AllJobs_Job_class, Crawl_Date, unique_id
        FROM sites_datas
        WHERE Site=%(site)s AND Crawl_Date =%(crawl_date)s
        Order By Company
    """
    df_main = pd.read_sql(sql, conn, params={
        'site': site, 'crawl_date': today_str})
    df_main.to_excel(main_writer, site_name, index=False)
    main_writer.save()
    return file_name


def combile_files(email=False):
    directory = 'IL-jobcrawl-data'
    today = datetime.date.today()
    today_str = today.strftime("%Y_%m_%d")
    # today_str = '2016_09_24'
    # if not email:
    main_excel_path = '{0}/{1}_Daily-List-Of-Competitor-Jobs.xlsx'.format(
        directory, today_str)
    try:
        os.remove(main_excel_path)
    except:
        pass
    for sheet_name in ['Drushim', 'AllJobs', 'JobMaster']:
        # sheet_name = 'Drushim'
        file_name = sheet_name.lower().title()
        temp_each_site_excel_file_path = '{0}/{1}_{2}.xlsx'.format(
            directory, today_str, file_name
        )

        if not os.path.isfile(main_excel_path):
            wb = Workbook()
            wb.save(main_excel_path)

        main_book = load_workbook(main_excel_path)
        sheet_names = main_book.get_sheet_names()
        if len(sheet_names) > 1:
            try:
                std = main_book.get_sheet_by_name('Sheet')
                main_book.remove_sheet(std)
                main_book.save(main_excel_path)
            except:
                pass
        main_writer = pd.ExcelWriter(
            main_excel_path, engine='openpyxl')
        main_writer.book = main_book

        main_writer.sheets = dict(
            (ws.title, ws) for ws in main_book.worksheets)
        unsorted_xls_df = pd.read_excel(
            temp_each_site_excel_file_path)
        sorted_xls = unsorted_xls_df.sort_values(by='Company')
        sorted_xls = sorted_xls.drop_duplicates()

        sorted_xls.to_excel(main_writer, sheet_name, index=False)
        main_writer.save()

    if email:
        # send email for total site data
        file_name = '{}_Daily-List-Of-Competitor-Jobs.xlsx'.format(
            today_str)
        body = "Please find the attachment for {}".format(file_name)

        send_email(directory=directory, file_name=file_name, body=body)

today = datetime.date.today()
# today = datetime.date(2016, 10, 05)
today_str = today.strftime("%Y_%m_%d")


def clean_residual_data():
    base_path = os.path.dirname(os.path.realpath(__file__))
    data_path = '{0}/{1}'.format(base_path, 'IL-jobcrawl-data')
    client_path = '{0}/{1}'.format(
        base_path, 'daily_competitor_client_changes')
    logs_path = '{0}/{1}'.format(base_path, 'logs')

    # Generate range of dates for a week
    date_range = []
    for i in range(10):
        item = (today - datetime.timedelta(days=i)).strftime("%Y_%m_%d")
        date_range.append(item)

    for each_dir in [data_path, client_path, logs_path]:
        for each_file in os.listdir(each_dir):
            new_file = each_file.replace('scrapy_log_output_', '')
            file_date = new_file[:10]
            if file_date not in date_range:
                os.remove("{0}/{1}".format(each_dir, each_file))


if __name__ == '__main__':
    # combile_files(email=True)
    # generate_excel("Alljobs")
    # generate_excel("Jobmaster")
    # generate_excel("Drushim")
    clean_residual_data()
