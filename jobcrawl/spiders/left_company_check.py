import sys
import codecs
import scrapy
import locale
import time
import os
from xlrd import open_workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

from scrapy import signals
import pandas as pd
import datetime
from scrapy.xlib.pydispatch import dispatcher

from jobcrawl.mailer import send_email  # , send_plain_email
from jobcrawl.clientchanges import ClientChanges
from excel_gen import generate_excel

today = datetime.date.today()
# today = datetime.date.today() - datetime.timedelta(days=1)
today_str = today.strftime("%Y_%m_%d")


class LeftCompany(scrapy.Spider):
    """ Spider to verify removed companies """
    name = "left"
    allowed_domains = [
        "jobmaster.co.il", "drushim.co.il", "alljobs.co.il",
        "jobnet.co.il"
    ]

    start_urls = []
    excel_dir = 'daily_competitor_client_changes'
    excel_path = '{}/{}_Daily-Competitor-Client-Change.xlsx'.format(
        excel_dir, today_str)

    def __init__(self):

        dispatcher.connect(self.spider_closed, signals.spider_closed)
        sys.stdout = codecs.getwriter(
            locale.getpreferredencoding())(sys.stdout)
        reload(sys)
        sys.setdefaultencoding('utf-8')

        # prepare clientchanges report
        self.c = ClientChanges()
        self.c.start()

        # self.wb = load_workbook(self.excel_path)
        # self.left_sheet = self.wb.get_sheet_by_name('Companies_That_left')
        # self.new_sheet = self.wb.get_sheet_by_name('New_Companies')
        # self.new_sheet_write = self.wb.create_sheet('New', 0)
        # self.left_sheet_write = self.wb.create_sheet('Left', 1)

    # def start_requests(self):
    #     wb = open_workbook(self.excel_path)
    #     sheet = wb.sheet_by_name('Companies_That_left')

    #     for i in range(1, sheet.nrows):
    #         row = sheet.row_values(i)
    #         company = row[1]
    #         site = row[0]
    #         company_url = row[2]
    #         company_jobs = row[3]
    #         if company_url:
    #             company_detail = [site, company, company_url, company_jobs]
    #             time.sleep(0.5)
    #             yield scrapy.Request(
    #                 company_url, self.parse,
    #                 meta={'company_detail': company_detail, 'type': 'removed'},
    #                 dont_filter=True
    #             )
    #     # get links for new companies
    #     new_sheet = wb.sheet_by_name('New_Companies')

    #     new_header_row = new_sheet.row_values(0)
    #     font = Font(size=11, bold=True)
    #     self.new_sheet_write.append(new_header_row)
    #     for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    #         c = self.new_sheet_write[i + '1']
    #         c.font = font

    #     for i in range(1, new_sheet.nrows):
    #         new_row = new_sheet.row_values(i)
    #         new_company = new_row[1]
    #         new_site = new_row[0]
    #         new_company_url = new_row[2]
    #         new_company_jobs = new_row[3]
    #         if new_company_url:
    #             new_company_detail = [
    #                 new_site, new_company, new_company_url, new_company_jobs]
    #             time.sleep(0.5)
    #             yield scrapy.Request(
    #                 new_company_url, self.parse,
    #                 meta={'company_detail': new_company_detail, 'type': 'new'},
    #                 dont_filter=True
    #             )

    def parse(self, response):
        pass
        # company_detail = response.meta['company_detail']
        # if response.meta['type'] == 'removed':
        #     drushimob_div = response.xpath("//div[@class='jobCount']")
        #     alljobs_jobs_div = response.xpath("//div[@class='job-paging']")
        #     jobmaster_jobs = response.xpath(
        #         "//div[@class='CenterContent']/article")
        #     jobnet_jobs = response.xpath("//div[@typeof='JobPosting']")
        #     if (
        #         drushimob_div or alljobs_jobs_div or jobmaster_jobs or
        #         jobnet_jobs or '/Search/' in response.url
        #     ):
        #         self.left_sheet.append(company_detail)
        # else:
        #     # open alljobs
        #     company_site_url = ''
        #     if company_detail[0] == 'AllJobs':
        #         try:
        #             company_site_url = response.xpath(
        #                 "//div[@id='divTagCompanyCategory']/"
        #                 "following-sibling::div"
        #             )[0].xpath("./a/text()").extract_first()
        #         except:
        #             pass

        #     company_detail.extend([company_site_url, '', ''])
        #     self.new_sheet_write.append(company_detail)

    def spider_closed(self, spider):
        # self.wb.remove_sheet(self.new_sheet)
        # self.new_sheet_write.title = "New_Companies"
        # self.wb.save(self.excel_path)

        # new_companies_df = pd.read_excel(
        #     self.excel_path, sheetname='New_Companies')
        # new_companies_df = new_companies_df.sort_values(
        #     by=['Site', 'Num_Company_jobs', 'Company'],
        #     ascending=[True, False, True]
        # )
        # left_companies_df = pd.read_excel(
        #     self.excel_path, sheetname='Companies_That_left')
        # left_companies_df = left_companies_df.drop_duplicates(keep=False)
        # left_companies_df = left_companies_df.sort_values(
        #     by=['Site', 'Num_Company_jobs', 'Company'],
        #     ascending=[True, False, True]
        # )

        # writer = pd.ExcelWriter(self.excel_path, engine='openpyxl')
        # new_companies_df.to_excel(writer, 'New_Companies', index=False)
        # left_companies_df.to_excel(writer, 'Companies_That_left', index=False)
        # writer.save()
        # self.logger.info('Saved clientchange sheet.')

        # send email for competitior changes
        directory = 'daily_competitor_client_changes'
        file_name = '{}_Daily-Competitor-Client-Change.xlsx'.format(
            today_str)

        self.stats = self.c.get_stats()
        self.logger.info('Obtained stats')
        # condition = (
        #     self.stats['total_jobs']['Drushim'] and
        #     self.stats['total_jobs']['JobMaster'] and
        #     self.stats['total_jobs']['AllJobs'] and
        #     self.stats['total_jobs']['JobNet']
        # )
        # if not condition:
        #     # file corrupt
        #     send_plain_email(
        #         subject="IL Job site data corrupt",
        #         body="Data corrupt for {}. Please check.".format(
        #             today_str
        #         )
        #     )
        #     return

        body = """
Please find the attachment for {subject}.

--- New / Removed Companies per Site ---
Drushim : (new) {drushim_new}, (removed) {drushim_removed}
JobMaster : (new) {jobmaster_new}, (removed) {jobmaster_removed}
AllJobs : (new) {alljobs_new}, (removed) {alljobs_removed}
JobNet : (new) {jobnet_new}, (removed) {jobnet_removed}

--- New Companies ---
Drushim : {drushim_new}
JobMaster : {jobmaster_new}
AllJobs : {alljobs_new}
JobNet : {jobnet_new}

--- Removed Companies ---
Drushim : {drushim_removed}
JobMaster : {jobmaster_removed}
AllJobs : {alljobs_removed}
JobNet : {jobnet_removed}
        """.format(
            subject=file_name, drushim_new=self.stats['new']['Drushim'],
            drushim_removed=self.stats['removed']['Drushim'],
            jobmaster_new=self.stats['new']['JobMaster'],
            jobmaster_removed=self.stats['removed']['JobMaster'],
            alljobs_new=self.stats['new']['AllJobs'],
            alljobs_removed=self.stats['removed']['AllJobs'],
            jobnet_new=self.stats['new']['JobNet'],
            jobnet_removed=self.stats['removed']['JobNet']
        )

        send_email(directory=directory, file_name=file_name, body=body)
        self.logger.info('Client change email sent')
        # send an email for 3 excel attachments
        directory = "IL-jobcrawl-data"
        file_to_send = []
        for site in ['Drushim', 'Alljobs', 'Jobmaster', 'Jobnet']:
            file_name = '{}_{}.xlsx'.format(
                today_str, site)
            # check if the file is corrupt
            try:
                load_workbook('{}/{}'.format(directory, file_name))
                self.logger.info('{} File good'.format(site))
            except:
                self.logger.info('{} file corrupt, regenerationg'.format(site))
                # file is corrupt, generate from sql
                generate_excel(site)
                self.logger.info('{} File generation success'.format(site))
            file_to_send.append(file_name)

        subject = '{}_Daily-List-Of-Competitor-Jobs.xlsx'.format(
            file_to_send[0][:10])
        body = """
Please find the attachment for {subject}.

--- Jobs / Companies per Site ---
Drushim : (jobs) {drushim_jobs}, (companies) {drushim_companies}
JobMaster : (jobs) {jm_jobs}, (companies) {jm_companies}
AllJobs : (jobs) {alljobs_jobs}, (companies) {alljobs_companies}
JobNet : (jobs) {jobnet_jobs}, (companies) {jobnet_companies}

--- Jobs per Site ---
Drushim : {drushim_jobs} jobs
JobMaster : {jm_jobs} jobs
AllJobs : {alljobs_jobs} jobs
JobNet : {jobnet_jobs} jobs

--- Companies per Site ---
Drushim : {drushim_companies} companies
JobMaster : {jm_companies} companies
AllJobs : {alljobs_companies} companies
JobNet : {jobnet_companies} companies
        """.format(
            subject=subject, drushim_jobs=self.stats['total_jobs']['Drushim'],
            drushim_companies=self.stats['total_companies']['Drushim'],
            jm_jobs=self.stats['total_jobs']['JobMaster'],
            jm_companies=self.stats['total_companies']['JobMaster'],
            alljobs_jobs=self.stats['total_jobs']['AllJobs'],
            alljobs_companies=self.stats['total_companies']['AllJobs'],
            jobnet_jobs=self.stats['total_jobs']['JobNet'],
            jobnet_companies=self.stats['total_companies']['JobNet']
        )

        send_email(
            directory=directory, file_name=file_to_send, body=body, multi=True)
        self.logger.info('Job list email sent')
        # delete old files
        self.clean_residual_data()
        self.logger.info('Residual data cleaned')
        self.logger.info('All done')

    def clean_residual_data(self):
        base_path = os.path.join(os.path.dirname(
            os.path.realpath(__file__)), os.pardir)
        base_path = os.path.abspath(os.path.join(base_path, os.pardir))

        data_path = '{0}/{1}'.format(base_path, 'IL-jobcrawl-data')
        client_path = '{0}/{1}'.format(
            base_path, 'daily_competitor_client_changes')
        logs_path = '{0}/{1}'.format(base_path, 'logs')

        # Generate range of dates for 10 days
        date_range = []
        for i in range(10):
            item = (today - datetime.timedelta(days=i)).strftime("%Y_%m_%d")
            date_range.append(item)

        month_range = []
        for i in range(30):
            item = (today - datetime.timedelta(days=i)).strftime("%d/%m/%Y")
            month_range.append(item)

        for each_dir in [data_path, client_path, logs_path]:
            for each_file in os.listdir(each_dir):
                new_file = each_file.replace('scrapy_log_output_', '')
                file_date = new_file[:10].replace('-', '_')
                if file_date not in date_range:
                    file_path = os.path.join(each_dir, each_file)
                    os.remove(file_path)
        self.c.clean_residual_database(month_range)
